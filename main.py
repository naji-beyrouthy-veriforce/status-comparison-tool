"""
Dynamics 365 vs SafeContractor Status Comparison
Automates ID extraction and status comparison reporting

Main business logic module - handles ID extraction and comparison generation.
Uses centralized configuration from config.py and utilities from utils.py
"""

import pandas as pd
import warnings
import sys
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure stdout encoding for Windows console compatibility
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass  # Python < 3.7 or already configured

# Suppress openpyxl style warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Import configuration
from src.config import (
    OUTPUT_DIR,
    DYNAMICS_DIR,
    REDASH_DIR,
    QUERY_IDS_DIR,
    D365_PATTERNS,
    SC_PATTERNS,
    D365_FILES,
    SC_FILES,
    REPORT_TYPES,
    MAX_FILE_SAVE_RETRIES,
    FILE_SAVE_RETRY_DELAY_SECONDS,
    CLIENT_STATUS_COLUMN,
    CASE_COLUMN_REPORT_TYPES,
    REDASH_API_KEY,
    Messages,
    setup_logging,
    get_dated_comparison_dir,
)

# Setup logging
logger = setup_logging("comparison_tool", console_output=True, file_output=True)

# Import utility functions
from src.utils import (
    clean_uuid,
    format_ids_for_sql,
    find_column_by_keywords,
    find_file_by_pattern,
    validate_file_format,
    validate_dataframe,
    apply_header_formatting,
    safe_read_excel,
    validate_uuid_data,
    check_file_accessibility,
    find_sc_status_column,
)

# Import email report generation
try:
    from src.email_report import generate_email_report as generate_report
except ImportError:
    logger.warning("Could not import email report generator")
    generate_report = None

# Import Redash API (optional - only needed for automated mode)
try:
    from src.redash_api import run_all_redash_queries
except ImportError:
    logger.warning("Could not import Redash API module")
    run_all_redash_queries = None


def extract_and_save_ids():
    """
    Step 1: Extract IDs from D365 files and save SQL-ready lists
    """
    logger.info("=" * 70)
    logger.info("STEP 1: EXTRACTING IDs FROM D365 FILES")
    logger.info("=" * 70)
    print("\n" + "=" * 70)
    print("STEP 1: EXTRACTING IDs FROM D365 FILES")
    print("=" * 70)

    # Process Accreditation and WCB only (Client doesn't need ID extraction)
    for report_type in ["accreditation", "wcb"]:
        logger.info(f"Processing {report_type} for ID extraction")
        print(Messages.processing(report_type))

        # Find D365 file by pattern in dynamics subdirectory
        file_path = find_file_by_pattern(DYNAMICS_DIR, D365_PATTERNS[report_type], "d365")
        if not file_path:
            # Try without suffix requirement
            file_path = find_file_by_pattern(DYNAMICS_DIR, D365_PATTERNS[report_type])

        if not file_path:
            logger.warning(f"No D365 {report_type} file found")
            print(Messages.warning(Messages.FILE_NOT_FOUND.format(report_type=report_type)))
            print(f"     {Messages.LOOKING_FOR.format(patterns=D365_PATTERNS[report_type])}")
            continue

        # Validate file format
        is_valid, error_msg, suggested_fix = validate_file_format(file_path)
        if not is_valid:
            logger.error(f"File validation failed for {file_path.name}: {error_msg}")
            print(Messages.error(error_msg))
            if suggested_fix:
                print(Messages.suggestion(suggested_fix))
            continue

        # Read file with enhanced error handling
        df, error_msg, suggested_fix = safe_read_excel(file_path)
        if error_msg:
            logger.error(f"Failed to read {file_path.name}: {error_msg}")
            print(Messages.error(f"reading {file_path.name}: {error_msg}"))
            if suggested_fix:
                print(Messages.suggestion(suggested_fix))
            continue

        logger.info(f"Successfully read {len(df)} rows from {file_path.name}")
        print(Messages.success(Messages.READ_ROWS.format(count=len(df), filename=file_path.name)))

        # Validate DataFrame structure
        is_valid, error_msg, suggested_fix = validate_dataframe(
            df, file_path.name, [("global", "alcumus", "id")]
        )
        if not is_valid:
            print(Messages.error(error_msg))
            if suggested_fix:
                print(Messages.suggestion(suggested_fix))
            continue

        # Find Global Alcumus Id column
        id_col = find_column_by_keywords(df.columns, ("global", "alcumus", "id"))

        if not id_col:
            logger.error(f"Global Alcumus ID column not found in {file_path.name}")
            print(Messages.error(Messages.COLUMN_NOT_FOUND))
            available = ", ".join([f"'{col}'" for col in df.columns])
            print(f"     {Messages.AVAILABLE_COLUMNS.format(columns=available)}")
            print(Messages.suggestion(Messages.ENSURE_EXPORT))
            continue

        # Validate UUID data quality
        is_valid, msg, fix, uuid_stats = validate_uuid_data(df, id_col, file_path.name)
        if not is_valid:
            logger.error(f"UUID validation failed: {msg}")
            print(Messages.error(msg))
            if fix:
                print(Messages.suggestion(fix))
            continue

        # Show UUID statistics
        print(
            Messages.info(
                Messages.UUID_QUALITY.format(
                    valid=uuid_stats["valid_uuids"],
                    total=uuid_stats["total"],
                    null=uuid_stats["null"],
                    invalid=uuid_stats["invalid"],
                )
            )
        )

        # Extract and clean IDs using vectorized operation
        unique_ids = df[id_col].dropna().map(clean_uuid).dropna().unique()
        unique_ids = sorted(unique_ids)

        # Final validation
        if len(unique_ids) == 0:
            logger.error(f"No valid UUIDs extracted from {id_col} column")
            print(Messages.error(Messages.NO_VALID_UUIDS.format(column=id_col)))
            sample_values = df[id_col].head(3).tolist()
            print(f"     {Messages.SAMPLE_VALUES.format(values=sample_values)}")
            print(Messages.suggestion(Messages.CHECK_COLUMN))
            continue

        logger.info(f"Extracted {len(unique_ids)} unique IDs from {report_type}")
        print(Messages.success(Messages.EXTRACTED_IDS.format(count=len(unique_ids))))
        print(Messages.success(Messages.USING_FRESH_IDS))

        # Format for SQL
        sql_formatted = format_ids_for_sql(unique_ids)

        # Save to file in query_ids subfolder (create parent dirs if needed)
        QUERY_IDS_DIR.mkdir(parents=True, exist_ok=True)

        output_file = QUERY_IDS_DIR / f"{report_type}_ids.sql.txt"

        with open(output_file, "w") as f:
            f.write(sql_formatted)

        logger.info(f"Saved {len(unique_ids)} IDs to {output_file.name}")
        print(Messages.success(Messages.SAVED_TO.format(filename=output_file.name)))

        # Show preview
        lines = sql_formatted.split("\n")
        print(f"  {Messages.PREVIEW_HEADER}")
        for line in lines[:5]:
            print(f"    {line}")
        if len(lines) > 5:
            print(f"    {Messages.AND_MORE.format(count=len(lines) - 5)}")

    logger.info("ID extraction completed successfully")
    print("\n" + "=" * 70)
    print("✅ ID EXTRACTION COMPLETED!")
    print("=" * 70 + "\n")


def create_comparison_excel(report_type, df_d365, df_sc):
    """
    Create comparison Excel file with SC and D365 sheets.
    
    Logs all major operations and data quality metrics.

    COMPARISON LOGIC:
    ========================================================
    1. Two sheets created: "SC" sheet and "D365" sheet
    2. Each sheet gets XLOOKUP formulas to pull status from the other sheet
    3. "Is it the same?" column compares statuses:
       - Accreditation/WCB: SC 'status' column vs D365 Status
       - Client: SC 'case' column (which IS the status) vs D365 Status
    4. Column placement varies by report type for readability:
       - Client: Inserts comparison columns after 'case' column
       - Accreditation/WCB: Appends comparison columns at the end

    ⚠️ CRITICAL - DO NOT MODIFY CLIENT COMPARISON LOGIC:
    The 'case' column in Client SafeContractor data IS the status column.
    This is a business requirement from the Redash query structure.
    Comparing 'case' vs D365 Status is CORRECT, not a bug!

    Args:
        report_type: Type of report (accreditation, wcb, client, or critical_document)
        df_d365: Dynamics 365 DataFrame
        df_sc: SafeContractor DataFrame

    Returns:
        Path to created Excel file, or None if creation failed
    """
    logger.info(f"Creating comparison for {report_type}: D365={len(df_d365)} rows, SC={len(df_sc)} rows")
    print(
        Messages.info(
            Messages.CREATING_COMPARISON.format(report_type=report_type)
        )
    )
    print(
        f"     {Messages.ROW_COUNTS.format(d365_count=len(df_d365), sc_count=len(df_sc))}"
    )

    # Find D365 columns using helper function
    id_col_d365 = find_column_by_keywords(df_d365.columns, ("global", "alcumus", "id"))
    status_col_d365 = find_column_by_keywords(df_d365.columns, ("status", "reason"))

    if not id_col_d365 or not status_col_d365:
        logger.error(f"Missing required D365 columns - ID: {id_col_d365}, Status: {status_col_d365}")
        print(f"     {Messages.MISSING_COLUMNS}")
        print(f"        ID column: {id_col_d365}")
        print(f"        Status column: {status_col_d365}")
        return None

    print(f"     {Messages.COLUMN_INFO.format(id_col=id_col_d365)}")
    print(f"     {Messages.STATUS_INFO.format(status_col=status_col_d365)}")

    # Find SC columns intelligently
    id_col_sc = (
        find_column_by_keywords(df_sc.columns, ("global", "alcumus", "id"), ("id", "alcumus"))
        or df_sc.columns[0]
    )

    # Find status column in SC data using the centralized helper function
    status_col_sc = find_sc_status_column(df_sc, id_col_sc, report_type)

    if not status_col_sc:
        print(f"     {Messages.STATUS_COLUMN_MISSING}")
        print(f"        Available columns: {list(df_sc.columns)}")
        return None

    print(f"     SC ID column: '{id_col_sc}'")
    print(f"     SC Status column: '{status_col_sc}'")

    # Clean IDs in both dataframes
    df_d365["clean_id"] = df_d365[id_col_d365].apply(clean_uuid)
    df_sc["clean_id"] = df_sc[id_col_sc].apply(clean_uuid)

    # Verify cleaned IDs
    d365_clean_count = df_d365["clean_id"].notna().sum()
    sc_clean_count = df_sc["clean_id"].notna().sum()
    print(f"     D365 cleaned IDs: {d365_clean_count}/{len(df_d365)}")
    print(f"     SC cleaned IDs: {sc_clean_count}/{len(df_sc)}")

    # Check for matches
    common_ids = set(df_d365["clean_id"].dropna()) & set(df_sc["clean_id"].dropna())
    logger.info(f"Found {len(common_ids)} common IDs between D365 and SC for {report_type}")
    print(f"     Common IDs found: {len(common_ids)}")

    if len(common_ids) == 0:
        logger.warning(f"No matching IDs found between D365 and SC for {report_type}")
        logger.debug(f"Sample D365 IDs: {list(df_d365['clean_id'].dropna()[:3])}")
        logger.debug(f"Sample SC IDs: {list(df_sc['clean_id'].dropna()[:3])}")
        print(f"     ⚠ WARNING: No matching IDs found between D365 and SC!")
        print(f"     Sample D365 IDs: {list(df_d365['clean_id'].dropna()[:3])}")
        print(f"     Sample SC IDs: {list(df_sc['clean_id'].dropna()[:3])}")

    # ============================================================================
    # CRITICAL BUSINESS LOGIC - DO NOT MODIFY WITHOUT UNDERSTANDING:
    # ============================================================================
    # For CLIENT reports from SafeContractor Redash query:
    #   - The 'case' column IS the status column for client-specific global IDs
    #   - This is NOT the same as a regular 'status' column
    #   - Comparison logic MUST use 'case' column for client reports
    #
    # For ACCREDITATION/WCB reports:
    #   - The 'status' column is used normally
    #
    # This is the CORRECT behavior per business requirements.
    # ============================================================================

    # ===== CONVERT DATE COLUMNS TO PROPER DATETIME =====
    date_format = "YYYY-MM-DD HH:MM:SS"
    for df in (df_sc, df_d365):
        for col in df.columns:
            if col.lower().endswith("_at") or col.lower().endswith("_date"):
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")

    # ===== CREATE EXCEL FILE WITH TWO SHEETS AND XLOOKUP FORMULAS =====
    wb = Workbook()
    wb.remove(wb.active)

    # ===== SC SHEET (CREATED FIRST) =====
    ws_sc = wb.create_sheet("SC")

    # Write SC data (preserve original column order)
    sc_data_cols = df_sc.drop(columns=["clean_id"]).columns
    for r_idx, row in enumerate(
        dataframe_to_rows(df_sc.drop(columns=["clean_id"]), index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws_sc.cell(row=r_idx, column=c_idx, value=value)

    # Apply date format to date columns in SC sheet
    for c_idx, col in enumerate(sc_data_cols, 1):
        if col.lower().endswith("_at") or col.lower().endswith("_date"):
            for r_idx in range(2, ws_sc.max_row + 1):
                ws_sc.cell(row=r_idx, column=c_idx).number_format = date_format

    # Find SC ID and Status column positions
    sc_cols = list(df_sc.drop(columns=["clean_id"]).columns)
    sc_id_col_idx = sc_cols.index(id_col_sc) + 1
    sc_status_col_idx_orig = sc_cols.index(status_col_sc) + 1
    sc_id_col_letter = ws_sc.cell(1, sc_id_col_idx).column_letter
    sc_status_col_letter = ws_sc.cell(1, sc_status_col_idx_orig).column_letter

    # Determine where to insert new comparison columns based on report type
    sc_cols_lower = {col.lower(): idx for idx, col in enumerate(sc_cols, 1)}
    is_client = report_type.lower() in CASE_COLUMN_REPORT_TYPES

    # COLUMN PLACEMENT STRATEGY:
    # - Client/Critical Document reports: Insert after CLIENT_STATUS_COLUMN for better visibility
    # - Accreditation/WCB: Append at the end
    if is_client and CLIENT_STATUS_COLUMN.lower() in sc_cols_lower:
        # For client reports, insert after the CLIENT_STATUS_COLUMN
        insert_after_idx = sc_cols_lower[CLIENT_STATUS_COLUMN.lower()]

        # Insert two columns: "D365 Status" and "Is it the same?"
        ws_sc.insert_cols(insert_after_idx + 1, 2)

        # CRITICAL: Update status column position if it shifted due to insertion
        if sc_status_col_idx_orig > insert_after_idx:
            sc_status_col_idx_orig += 2
            sc_status_col_letter = ws_sc.cell(1, sc_status_col_idx_orig).column_letter

        # Set headers for inserted columns
        ws_sc.cell(1, insert_after_idx + 1, "D365 Status")
        ws_sc.cell(1, insert_after_idx + 2, "Is it the same?")
    else:
        # For Accreditation and WCB, add columns at the end
        insert_after_idx = len(sc_cols)

        # Set headers at the end
        ws_sc.cell(1, insert_after_idx + 1, "D365 Status")
        ws_sc.cell(1, insert_after_idx + 2, "Is it the same?")

    # Format specific headers (red fill, black bold text)
    apply_header_formatting(ws_sc)

    # Enable autofilter on headers
    ws_sc.auto_filter.ref = ws_sc.dimensions

    # ===== D365 SHEET (CREATED SECOND) =====
    ws_d365 = wb.create_sheet("D365")

    # Write D365 data (preserve original column order)
    d365_data_cols = df_d365.drop(columns=["clean_id"]).columns
    for r_idx, row in enumerate(
        dataframe_to_rows(df_d365.drop(columns=["clean_id"]), index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws_d365.cell(row=r_idx, column=c_idx, value=value)

    # Apply date format to date columns in D365 sheet
    for c_idx, col in enumerate(d365_data_cols, 1):
        if col.lower().endswith("_at") or col.lower().endswith("_date"):
            for r_idx in range(2, ws_d365.max_row + 1):
                ws_d365.cell(row=r_idx, column=c_idx).number_format = date_format

    # Find D365 ID and Status column positions (1-indexed for Excel)
    d365_cols = list(df_d365.drop(columns=["clean_id"]).columns)
    d365_id_col_idx = d365_cols.index(id_col_d365) + 1
    d365_status_col_idx = d365_cols.index(status_col_d365) + 1
    d365_id_col_letter = ws_d365.cell(1, d365_id_col_idx).column_letter

    # Add SC Status column with XLOOKUP formula
    sc_status_col_idx = len(d365_cols) + 1
    ws_d365.cell(1, sc_status_col_idx, "SC Status")

    # Add "Is it the same?" column
    is_same_col_idx = sc_status_col_idx + 1
    ws_d365.cell(1, is_same_col_idx, "Is it the same?")

    # Format specific headers (red fill, black bold text)
    apply_header_formatting(ws_d365)

    # Enable autofilter on headers
    ws_d365.auto_filter.ref = ws_d365.dimensions

    # ============================================================================
    # CRITICAL: Determine which column to use for status lookups
    # ============================================================================
    # CLIENT/CRITICAL DOCUMENT REPORTS: 'case' column IS the status column
    #   - The SafeContractor Redash queries for these report types
    #     return status in the 'case' column
    #   - Must compare 'case' vs D365 Status for accurate comparison
    #
    # ACCREDITATION/WCB REPORTS: Regular 'status' column is used
    #   - Standard status comparison logic applies
    #
    # DO NOT "FIX" THIS - This is the CORRECT implementation!
    # ============================================================================
    comparison_col_letter = sc_status_col_letter
    if report_type.lower() in CASE_COLUMN_REPORT_TYPES:
        case_col_idx = next(
            (idx for idx, col in enumerate(sc_cols, 1) if col.lower() == CLIENT_STATUS_COLUMN.lower()), None
        )
        if case_col_idx:
            # Adjust if columns were inserted before the case column
            if case_col_idx > insert_after_idx:
                case_col_idx += 2
            comparison_col_letter = ws_sc.cell(1, case_col_idx).column_letter

    # Log comparison logic for verification
    comparison_col_name = CLIENT_STATUS_COLUMN if report_type.lower() in CASE_COLUMN_REPORT_TYPES else status_col_sc
    print(f"     Comparison Logic:")
    print(f"       - D365 Sheet: Comparing D365 '{status_col_d365}' vs SC '{comparison_col_name}'")
    print(f"       - SC Sheet: Comparing SC '{comparison_col_name}' vs D365 Status")

    # Add XLOOKUP formulas for D365 sheet (row 2 onwards)
    # Cache column letters for better performance
    d365_status_col_letter = ws_d365.cell(1, d365_status_col_idx).column_letter
    sc_status_lookup_col_letter = ws_d365.cell(1, sc_status_col_idx).column_letter

    for row_idx in range(2, len(df_d365) + 2):
        # XLOOKUP with _xlfn prefix and entire column references
        # FIX: Use comparison_col_letter instead of sc_status_col_letter for correct column
        xlookup_formula = f'=_xlfn.XLOOKUP({d365_id_col_letter}{row_idx},SC!{sc_id_col_letter}:{sc_id_col_letter},SC!{comparison_col_letter}:{comparison_col_letter},"Not found",0)'
        ws_d365.cell(row_idx, sc_status_col_idx, xlookup_formula)

        # Is it the same? formula
        ws_d365.cell(
            row_idx,
            is_same_col_idx,
            f"={d365_status_col_letter}{row_idx}={sc_status_lookup_col_letter}{row_idx}",
        )

    # Add XLOOKUP formulas for SC sheet (row 2 onwards)
    d365_lookup_col_letter = ws_sc.cell(1, insert_after_idx + 1).column_letter

    for row_idx in range(2, len(df_sc) + 2):
        # XLOOKUP with _xlfn prefix and entire column references
        xlookup_formula = f'=_xlfn.XLOOKUP({sc_id_col_letter}{row_idx},D365!{d365_id_col_letter}:{d365_id_col_letter},D365!{d365_status_col_letter}:{d365_status_col_letter},"Not found",0)'
        ws_sc.cell(row_idx, insert_after_idx + 1, xlookup_formula)

        # Is it the same? formula - compares appropriate status column
        ws_sc.cell(
            row_idx,
            insert_after_idx + 2,
            f"={comparison_col_letter}{row_idx}={d365_lookup_col_letter}{row_idx}",
        )

    # Get dated comparison directory (e.g., output/comparison_2026-02-18/)
    comparison_dir = get_dated_comparison_dir()
    
    # Create the directory if it doesn't exist
    comparison_dir.mkdir(parents=True, exist_ok=True)
    
    # Save file with retry logic for locked files
    output_file = comparison_dir / f"{report_type.title()}_Comparison.xlsx"

    # Check if file is writable before attempting save
    if output_file.exists():
        is_accessible, msg, fix = check_file_accessibility(output_file, mode="write")
        if not is_accessible:
            print(Messages.warning(msg))
            if fix:
                print(Messages.suggestion(fix))
            # Try with timestamp immediately

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = comparison_dir / f"{report_type.title()}_Comparison_{timestamp}.xlsx"
            print(f"     💾 Saving as: {output_file.name}")

    # Try to save with retries
    for attempt in range(MAX_FILE_SAVE_RETRIES):
        try:
            wb.save(output_file)
            print(Messages.success(f"Successfully saved: {output_file.name}"))
            return output_file

        except PermissionError:
            if attempt < MAX_FILE_SAVE_RETRIES - 1:
                # Try again after a short delay
                logger.warning(f"File locked (attempt {attempt + 1}/{MAX_FILE_SAVE_RETRIES}): {output_file.name}")
                print(
                    Messages.warning(
                        Messages.FILE_LOCKED.format(
                            attempt=attempt + 1, max_attempts=MAX_FILE_SAVE_RETRIES
                        )
                    )
                )
                print(Messages.suggestion(Messages.CLOSE_FILE.format(filename=output_file.name)))

                time.sleep(FILE_SAVE_RETRY_DELAY_SECONDS)
            else:
                # Final attempt failed - save with timestamp
                logger.error(f"File still locked after {MAX_FILE_SAVE_RETRIES} attempts: {output_file.name}")
                print(
                    Messages.error(
                        Messages.FILE_STILL_LOCKED.format(max_attempts=MAX_FILE_SAVE_RETRIES)
                    )
                )

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = comparison_dir / f"{report_type.title()}_Comparison_{timestamp}.xlsx"

                try:
                    wb.save(backup_file)
                    logger.info(f"Saved with timestamp: {backup_file.name}")
                    print(f"     ✅ Saved with timestamp: {backup_file.name}")
                    print(
                        Messages.suggestion(
                            Messages.REMEMBER_TO_CLOSE.format(filename=output_file.name)
                        )
                    )
                    return backup_file
                except Exception as e:
                    logger.critical(f"Critical save error for {output_file.name}: {str(e)}")
                    print(Messages.error(Messages.CRITICAL_SAVE_ERROR))
                    print(f"     Error: {e}")
                    print(
                        Messages.suggestion(
                            Messages.CHECK_DISK_SPACE.format(directory=OUTPUT_DIR)
                        )
                    )
                    raise

        except Exception as e:
            error_type = type(e).__name__
            logger.error(f"Unexpected error saving file: {error_type} - {str(e)}")
            print(
                Messages.error(
                    Messages.UNEXPECTED_ERROR.format(error_type=error_type)
                )
            )
            print(f"     Details: {str(e)}")
            print(Messages.suggestion(Messages.CHECK_WRITABLE))
            raise

    return output_file


def generate_comparisons():
    """
    Step 2: Generate comparison Excel files
    """
    logger.info("=" * 70)
    logger.info("STEP 2: GENERATING COMPARISON FILES")
    logger.info("=" * 70)
    print("\n" + "=" * 70)
    print("STEP 2: GENERATING COMPARISON FILES")
    print("=" * 70)

    success_count = 0

    for report_type in REPORT_TYPES:
        logger.info(f"Processing comparison for {report_type}")
        print(Messages.processing(report_type))

        # Check if files exist in subdirectories
        d365_file = DYNAMICS_DIR / D365_FILES[report_type]
        sc_file = REDASH_DIR / SC_FILES[report_type]

        if not d365_file.exists():
            logger.warning(f"D365 file not found: {d365_file.name}")
            print(Messages.warning(f"{d365_file.name} not found, skipping..."))
            continue

        if not sc_file.exists():
            logger.warning(f"SC file not found: {sc_file.name}")
            print(Messages.warning(f"{sc_file.name} not found, skipping..."))
            continue

        # Validate file formats
        is_valid_d365, error_msg_d365, fix_d365 = validate_file_format(d365_file)
        if not is_valid_d365:
            print(Messages.error(error_msg_d365))
            if fix_d365:
                print(Messages.suggestion(fix_d365))
            continue

        is_valid_sc, error_msg_sc, fix_sc = validate_file_format(sc_file)
        if not is_valid_sc:
            print(Messages.error(error_msg_sc))
            if fix_sc:
                print(Messages.suggestion(fix_sc))
            continue

        # Read files with error handling
        df_d365, error_d365, fix_d365 = safe_read_excel(d365_file)
        if error_d365:
            print(Messages.error(f"reading {d365_file.name}: {error_d365}"))
            if fix_d365:
                print(Messages.suggestion(fix_d365))
            continue

        df_sc, error_sc, fix_sc = safe_read_excel(sc_file)
        if error_sc:
            print(Messages.error(f"reading {sc_file.name}: {error_sc}"))
            if fix_sc:
                print(Messages.suggestion(fix_sc))
            continue

        print(Messages.success(Messages.READ_D365.format(count=len(df_d365))))
        print(Messages.success(Messages.READ_SC.format(count=len(df_sc))))

        # Validate DataFrame structures
        required_d365_cols = [("global", "alcumus", "id"), ("status",)]
        is_valid, error_msg, fix = validate_dataframe(df_d365, d365_file.name, required_d365_cols)
        if not is_valid:
            print(Messages.error(error_msg))
            if fix:
                print(Messages.suggestion(fix))
            continue

        required_sc_cols = [("id",)]  # SC file should have at least an ID column
        is_valid, error_msg, fix = validate_dataframe(df_sc, sc_file.name, required_sc_cols)
        if not is_valid:
            print(Messages.error(error_msg))
            if fix:
                print(Messages.suggestion(fix))
            continue

        try:
            # Create comparison
            output_file = create_comparison_excel(report_type, df_d365, df_sc)

            if output_file:
                logger.info(f"Successfully created comparison file: {output_file.name}")
                print(Messages.success(Messages.CREATED_FILE.format(filename=output_file.name)))
                success_count += 1
            else:
                logger.error(f"Failed to create comparison for {report_type}")
                print(Messages.error(Messages.FAILED_COMPARISON))

        except Exception as e:
            logger.exception(f"Error processing {report_type}: {str(e)}")
            print(f"{Messages.ERROR} Error processing {report_type}: {e}")
            continue

    logger.info(f"Comparison generation completed: {success_count} files created")
    print("\n" + "=" * 70)
    if success_count > 0:
        # Get the comparison directory name
        comparison_dir = get_dated_comparison_dir()
        folder_name = comparison_dir.name
        
        print(f"SUCCESS! Created {success_count} comparison file(s) in {folder_name}/")
        print(f"         Location: {comparison_dir}")
        
        # Automatically generate email report
        print("\n" + "=" * 70)
        print("GENERATING EMAIL REPORT...")
        print("=" * 70)
        
        if generate_report:
            try:
                logger.info("Starting automatic email report generation")
                generate_report()
                logger.info("Email report generation completed")
            except Exception as e:
                logger.error(f"Failed to generate email report: {e}")
                print(Messages.warning(f"Could not generate email report: {e}"))
                print("     You can run it manually with: python generate_email_report.py")
        else:
            logger.warning("Email report generator not available")
            print(Messages.warning("Email report generator not available"))
    else:
        print("No comparison files were created. Check file locations.")
    print("=" * 70 + "\n")


def run_automated_workflow():
    """
    Automated workflow: Extract IDs → Redash queries → Generate comparisons.
    All steps run automatically — no manual Redash intervention needed.
    
    Requires REDASH_API_KEY environment variable to be set.
    """
    logger.info("Starting Automated Workflow")
    print("\n" + "=" * 70)
    print("AUTOMATED WORKFLOW: D365 vs SafeContractor Comparison")
    print("=" * 70)

    # Step 1: Extract IDs from D365 files (existing function)
    print("\n📋 STEP 1/3: Extracting IDs from D365 files...")
    extract_and_save_ids()

    # Step 2: Run Redash queries automatically (new)
    print("\n📡 STEP 2/3: Running Redash queries...")
    try:
        redash_results = run_all_redash_queries()
    except Exception as e:
        logger.error(f"Redash automation failed: {e}")
        print(f"\n{Messages.ERROR} Redash automation failed: {e}")
        print("\nFalling back to manual workflow:")
        print("1. Copy IDs from output/query_ids/*.sql.txt files")
        print("2. Paste into Redash queries and execute")
        print("3. Download results and place in input/redash/")
        print("4. Run this script again")
        return

    if not redash_results:
        print(f"\n{Messages.ERROR} No Redash results downloaded. Cannot proceed.")
        return

    # Step 3: Generate comparisons (existing function)
    print("\n📊 STEP 3/3: Generating comparison files...")
    generate_comparisons()

    print("\n" + "=" * 70)
    print("✅ AUTOMATED WORKFLOW COMPLETE!")
    print("=" * 70 + "\n")


def main():
    """
    Main execution flow.
    - If REDASH_API_KEY is set → automated mode (extract → Redash → compare)
    - Otherwise → manual workflow (existing 3-step process)
    """
    if REDASH_API_KEY and run_all_redash_queries:
        # Automated mode
        logger.info("Starting Status Comparison Tool - Automated Mode")
        print("\n" + "=" * 70)
        print("DYNAMICS 365 vs SAFECONTRACTOR STATUS COMPARISON")
        print("Automated Mode (Redash API)")
        print("=" * 70)

        # Check if D365 files exist
        d365_files_exist = any(
            find_file_by_pattern(DYNAMICS_DIR, D365_PATTERNS[t]) is not None
            for t in REPORT_TYPES
        )

        if d365_files_exist:
            run_automated_workflow()
        else:
            print(f"\n{Messages.WARNING} No D365 files found in input/dynamics/")
            print("Upload D365 files first, then run again.")
    else:
        # Manual mode (original behavior)
        logger.info("Starting Status Comparison Tool - Manual Workflow Mode")
        print("\n" + "=" * 70)
        print("DYNAMICS 365 vs SAFECONTRACTOR STATUS COMPARISON")
        print("Manual Workflow Mode")
        print("=" * 70)

        sc_files_exist = any(
            find_file_by_pattern(REDASH_DIR, SC_PATTERNS[t]) is not None for t in REPORT_TYPES
        )

        if sc_files_exist:
            print(f"\n{Messages.ALL_FILES_FOUND}")
            generate_comparisons()
        else:
            print(f"\n{Messages.WARNING} SC files not found - Starting with ID extraction...")
            extract_and_save_ids()


if __name__ == "__main__":
    main()
